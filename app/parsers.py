"""Excel file parsing and data normalization."""

import pandas as pd
import numpy as np
import re
from typing import Tuple, Dict, Optional
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell


def normalize_and_rename_columns(df: pd.DataFrame, sheet_type: str) -> pd.DataFrame:
    """
    Clean and standardize column names for both Grades and Attendance sheets.
    Handles minor naming variations and formatting differences.
    
    Args:
        df: DataFrame to normalize
        sheet_type: "grades" or "attendance"
    
    Returns:
        DataFrame with normalized column names
    """
    # Create a copy to avoid modifying the original
    df = df.copy()
    
    # Normalize base names (lowercase, trimmed, remove dots/%/commas, normalize whitespace)
    def normalize_col_name(col_name):
        """Normalize a column name for matching."""
        if pd.isna(col_name):
            return ""
        # Convert to string, lowercase, remove special chars, normalize whitespace
        normalized = str(col_name).strip().lower()
        normalized = re.sub(r'[.,%#]', '', normalized)  # Remove dots, commas, %, #
        normalized = re.sub(r'\s+', ' ', normalized)  # Normalize whitespace
        return normalized.strip()
    
    original_columns = df.columns.tolist()
    
    if sheet_type == "grades":
        # Define target standard names and their possible variations
        target_mappings = {
            "Student#": ["student#", "student number", "student id", "studentid", "studentnum"],
            "Student Name": ["student name", "studentname", "name", "student"],
            "Program Name": ["program name", "programname", "program", "course name", "coursename"],
            "current overall Program Grade": [
                "program grade", "grade", "overall program grade", 
                "current overall program grade", "overall grade", "final grade",
                "current grade", "programgrade"
            ],
        }
    elif sheet_type == "attendance":
        target_mappings = {
            "Student#": ["student#", "student number", "student id", "studentid", "studentnum"],
            "Student Name": ["student name", "studentname", "name", "student"],
            "Scheduled Hours to Date": ["scheduled hours to date", "scheduled hours", "scheduledhours"],
            "Attended Hours to Date": ["attended hours to date", "attended hours", "attendedhours"],
            "Attended % to Date.": [
                "attended to date", "attended percent to date", "attended % to date",
                "attended% to date", "attendance percent", "attendance %", "attended pct"
            ],
            "Missed Hours to Date": ["missed hours to date", "missed hours", "missedhours"],
            "% Missed": ["% missed", "missed", "missed percent", "missed %", "missed%"],
            "Missed Minus Excused to date": [
                "missed minus excused to date", "missed minus excused", 
                "missed minus excused hours"
            ],
            "Campus Login URL": ["campus login url", "campus login", "login url", "campusurl"],
        }
    else:
        target_mappings = {}
    
    # Build actual rename mapping by matching normalized column names
    actual_rename = {}
    for orig_col in original_columns:
        normalized = normalize_col_name(orig_col)
        
        # Try to find a match in target_mappings
        for target_name, variations in target_mappings.items():
            if normalized in variations:
                # Only rename if target_name doesn't already exist (avoid duplicates)
                if target_name not in df.columns or target_name not in actual_rename.values():
                    actual_rename[orig_col] = target_name
                    print(f"DEBUG: Will rename '{orig_col}' -> '{target_name}' (normalized: '{normalized}')")
                break
    
    # Apply renaming
    if actual_rename:
        df = df.rename(columns=actual_rename)
        print(f"DEBUG: Renamed columns in {sheet_type} sheet: {actual_rename}")
        print(f"DEBUG: Columns after renaming: {list(df.columns)}")
        
        # DEBUG: After renaming, check the first row to see what data is in each column
        if len(df) > 0:
            print(f"DEBUG: First row data after renaming:")
            for col in df.columns:
                val = df.iloc[0][col]
                print(f"  {col}: {val} (type: {type(val).__name__})")
    else:
        print(f"WARNING: No columns were renamed in {sheet_type} sheet. Original columns: {list(df.columns)}")
    
    # Remove duplicate columns (keep first occurrence)
    if df.columns.duplicated().any():
        print(f"WARNING: Found duplicate columns in {sheet_type} sheet: {df.columns[df.columns.duplicated()].tolist()}")
        df = df.loc[:, ~df.columns.duplicated(keep='first')]
        print(f"DEBUG: After removing duplicates: {list(df.columns)}")
    
    # Ensure required columns exist - if not found, try to create them or use closest match
    try:
        if sheet_type == "grades":
            # Handle Student Name - be more flexible in finding it
            if "Student Name" not in df.columns:
                print(f"DEBUG: 'Student Name' not found in grades columns: {list(df.columns)}")
                # Try to find any column that might be student name (more flexible matching)
                found_student_name = False
                for col in df.columns:
                    col_normalized = normalize_col_name(col)
                    # Check if it matches any variation
                    if col_normalized in ["student name", "studentname", "name", "student"]:
                        if col != "Student Name":  # Avoid renaming if already correct
                            df = df.rename(columns={col: "Student Name"})
                            print(f"DEBUG: Found and renamed '{col}' to 'Student Name'")
                            found_student_name = True
                            break
                
                if not found_student_name:
                    # If still not found, create a placeholder column
                    print("WARNING: 'Student Name' column not found in grades, creating placeholder")
                    if len(df) > 0:
                        df["Student Name"] = "Unknown"
                    else:
                        df["Student Name"] = pd.Series([], dtype=str)
            
            # Verify Student Name exists now
            if "Student Name" not in df.columns:
                raise ValueError(f"Failed to create 'Student Name' column in grades sheet. DataFrame shape: {df.shape}, columns: {list(df.columns)}")
            
            # Handle Student# - make it optional, create if missing
            if "Student#" not in df.columns:
                print("WARNING: 'Student#' column not found in grades sheet, will create sequential IDs")
                # Create sequential Student# based on index
                if len(df) > 0:
                    df["Student#"] = range(1, len(df) + 1)
                    print(f"DEBUG: Created Student# column with {len(df)} sequential IDs")
                else:
                    df["Student#"] = pd.Series([], dtype=int)
        
        elif sheet_type == "attendance":
            # Handle Student Name - be more flexible in finding it
            if "Student Name" not in df.columns:
                print(f"DEBUG: 'Student Name' not found in attendance columns: {list(df.columns)}")
                # Try to find any column that might be student name (more flexible matching)
                found_student_name = False
                for col in df.columns:
                    col_normalized = normalize_col_name(col)
                    # Check if it matches any variation
                    if col_normalized in ["student name", "studentname", "name", "student"]:
                        if col != "Student Name":  # Avoid renaming if already correct
                            df = df.rename(columns={col: "Student Name"})
                            print(f"DEBUG: Found and renamed '{col}' to 'Student Name'")
                            found_student_name = True
                            break
                
                if not found_student_name:
                    # If still not found, create a placeholder column
                    print("WARNING: 'Student Name' column not found in attendance, creating placeholder")
                    if len(df) > 0:
                        df["Student Name"] = "Unknown"
                    else:
                        df["Student Name"] = pd.Series([], dtype=str)
            
            # Verify Student Name exists now
            if "Student Name" not in df.columns:
                raise ValueError(f"Failed to create 'Student Name' column in attendance sheet. DataFrame shape: {df.shape}, columns: {list(df.columns)}")
            
            # Handle Student# - make it optional, create if missing
            if "Student#" not in df.columns:
                print("WARNING: 'Student#' column not found in attendance sheet, will create sequential IDs")
                # Create sequential Student# based on index
                if len(df) > 0:
                    df["Student#"] = range(1, len(df) + 1)
                    print(f"DEBUG: Created Student# column with {len(df)} sequential IDs")
                else:
                    df["Student#"] = pd.Series([], dtype=int)
        
        # Final verification
        print(f"DEBUG: Final columns after normalization ({sheet_type}): {list(df.columns)}")
        if "Student Name" not in df.columns:
            raise ValueError(f"'Student Name' column missing after normalization in {sheet_type} sheet. Columns: {list(df.columns)}")
        
    except Exception as e:
        print(f"ERROR in normalize_and_rename_columns for {sheet_type}: {e}")
        print(f"DataFrame shape: {df.shape}")
        print(f"DataFrame columns: {list(df.columns)}")
        raise
    
    return df


def safe_get_series(df: pd.DataFrame, column_name: str) -> pd.Series:
    """
    Safely extract a Series from a DataFrame, ensuring it's a Series not a DataFrame.
    
    Args:
        df: DataFrame to extract from
        column_name: Name of the column
    
    Returns:
        Series object
    """
    if column_name not in df.columns:
        raise KeyError(f"Column '{column_name}' not found in DataFrame")
    
    col_data = df[column_name]
    
    # If it's already a Series, return it
    if isinstance(col_data, pd.Series):
        return col_data
    
    # If it's a DataFrame (shouldn't happen, but handle it), extract first column
    if isinstance(col_data, pd.DataFrame):
        print(f"WARNING: Column '{column_name}' returned DataFrame instead of Series, using first column")
        return col_data.iloc[:, 0]
    
    # Otherwise, try to convert to Series
    return pd.Series(col_data, index=df.index)


def clean_student_id(df: pd.DataFrame) -> pd.DataFrame:
    """
    Standardize Student# column so merge keys match.
    Extracts digits only and converts to numeric.
    """
    if "Student#" in df.columns:
        # Safely get the Series
        student_id_series = safe_get_series(df, "Student#")
        # Convert to string first, then extract digits only
        df["Student#"] = (
            student_id_series
            .astype(str)
            .str.extract(r"(\d+)", expand=False)  # Extract digits only
        )
        df["Student#"] = pd.to_numeric(df["Student#"], errors="coerce")
    return df


def clean_attendance_df(df: pd.DataFrame) -> pd.DataFrame:
    """
    Clean, convert, and normalize attendance data.
    Removes total/summary rows, converts time columns, normalizes percentages.
    """
    df = df.copy()
    
    # Ensure Student Name exists before trying to filter
    if "Student Name" not in df.columns:
        print("WARNING: 'Student Name' column missing in clean_attendance_df, creating placeholder")
        if len(df) > 0:
            df["Student Name"] = "Unknown"
        else:
            df["Student Name"] = pd.Series([], dtype=str)
    
    # Remove total/summary rows
    try:
        if "Student Name" in df.columns and len(df) > 0:
            # Remove rows where Student Name contains "Total", "Summary", or is empty
            student_name_series = safe_get_series(df, "Student Name")
            mask = (
                ~student_name_series.astype(str).str.contains("Total", case=False, na=False) &
                ~student_name_series.astype(str).str.contains("Summary", case=False, na=False) &
                ~student_name_series.astype(str).str.strip().eq("") &
                student_name_series.notna()
            )
            df = df[mask].copy()
            print(f"DEBUG: Removed summary rows, remaining rows: {len(df)}")
    except Exception as e:
        print(f"WARNING: Error filtering summary rows: {e}. Continuing with all rows.")
    
    # Convert HH:MM columns to decimal hours
    time_columns = [
        "Scheduled Hours to Date",
        "Attended Hours to Date",
        "Missed Hours to Date",
        "Missed Minus Excused to date",
    ]
    
    for col in time_columns:
        if col in df.columns:
            df[col] = df[col].apply(to_hours)
    
    # Normalize % columns
    percentage_columns = ["Attended % to Date.", "% Missed"]
    
    for col in percentage_columns:
        if col in df.columns:
            df[col] = df[col].apply(normalize_pct)
    
    # Replace NaN and invalid values
    df = df.replace([np.inf, -np.inf, np.nan], 0)
    
    return df


def to_hours(value) -> float:
    """
    Convert time strings like '90:00' to numeric hours.
    Handles both string formats and numeric values.
    
    Args:
        value: String in format 'HH:MM', 'H:MM', or numeric value
    
    Returns:
        Decimal hours (e.g., '90:00' -> 90.0, '0:15' -> 0.25, 90.0 -> 90.0)
    """
    if isinstance(value, (int, float)):
        return float(value)
    
    if pd.isna(value) or value == '':
        return 0.0
    
    if isinstance(value, str) and ":" in value:
        try:
            hours, minutes = value.split(":")
            return float(hours) + float(minutes) / 60.0
        except (ValueError, TypeError):
            return 0.0
    
    # Try to parse as decimal number
    try:
        return float(value)
    except (ValueError, TypeError):
        return 0.0


def parse_duration(duration_str: str) -> float:
    """
    Parse duration string like '90:00' or '0:15' to decimal hours.
    (Alias for to_hours for backward compatibility)
    
    Args:
        duration_str: String in format 'HH:MM' or 'H:MM'
    
    Returns:
        Decimal hours (e.g., '90:00' -> 90.0, '0:15' -> 0.25)
    """
    return to_hours(duration_str)


def normalize_attendance_pct(x) -> float:
    """
    Normalize attendance percentage values.
    Handles both 0-1 decimals (e.g., 0.88) and 0-100 percentages (e.g., 88).
    
    Args:
        x: Value that might be in 0-1 range or 0-100 range
    
    Returns:
        Percentage in 0-100 range
    """
    if pd.isna(x) or x == '':
        return 0.0
    
    try:
        # Handle string values like "85%" or "0.85"
        if isinstance(x, str):
            val_str = str(x).strip().replace('%', '').strip()
            val = float(val_str)
        else:
            val = float(x)
        
        # If value is <= 1, assume it's a decimal (0-1 range) and multiply by 100
        # If value > 1, assume it's already a percentage (0-100 range)
        if val <= 1.0:
            return val * 100.0
        return val
    except (ValueError, TypeError):
        return 0.0


def normalize_pct(x) -> float:
    """
    Normalize percentage values (simpler version for direct column application).
    Handles both 0-1 decimals (e.g., 0.88) and 0-100 percentages (e.g., 88).
    
    Args:
        x: Value that might be in 0-1 range or 0-100 range
    
    Returns:
        Percentage in 0-100 range
    """
    # Handle NaN/None first
    if pd.isna(x) or x is None:
        return 0.0
    
    try:
        # Convert to float, handling string values
        if isinstance(x, str):
            # Remove % sign and whitespace
            val_str = str(x).strip().replace('%', '').strip()
            if not val_str or val_str == '':
                return 0.0
            val = float(val_str)
        else:
            val = float(x)
        
        # Check for invalid values
        if np.isnan(val) or np.isinf(val):
            return 0.0
        
        # If value is <= 1, assume it's a decimal (0-1 range) and multiply by 100
        # If value > 1, assume it's already a percentage (0-100 range)
        if val <= 1.0:
            return val * 100.0
        return val
    except (ValueError, TypeError) as e:
        print(f"DEBUG: normalize_pct error for value '{x}' (type: {type(x)}): {e}")
        return 0.0


def normalize_attendance_data(df: pd.DataFrame) -> pd.DataFrame:
    """
    Preprocessing step to automatically clean and normalize attendance data.
    
    This function:
    - Converts all "HH:MM" time strings into decimal hours (e.g., "89:45" -> 89.75)
    - Normalizes attendance percentages (e.g., 0.997 -> 99.7)
    - Replaces NaN or invalid values with 0
    
    Args:
        df: Raw attendance DataFrame
    
    Returns:
        Cleaned and normalized attendance DataFrame
    """
    # Make a copy to avoid modifying the original
    df = df.copy()
    
    # Debug: Print available columns before processing
    print(f"\n=== DEBUG: normalize_attendance_data - Available columns ===")
    print(f"Columns: {list(df.columns)}")
    
    # Convert HH:MM text columns to decimal hours
    time_columns = [
        "Scheduled Hours to Date",
        "Attended Hours to Date",
        "Missed Hours to Date",
        "Missed Minus Excused to date",
    ]
    
    for col in time_columns:
        if col in df.columns:
            print(f"DEBUG: Converting time column '{col}' - Sample values before: {df[col].head(3).tolist()}")
            df[col] = df[col].apply(to_hours)
            print(f"DEBUG: After conversion - Sample values: {df[col].head(3).tolist()}")
        else:
            # Try case-insensitive search
            for actual_col in df.columns:
                if col.lower().strip() == actual_col.lower().strip():
                    print(f"DEBUG: Found column '{actual_col}' (case variation of '{col}')")
                    print(f"DEBUG: Sample values before: {df[actual_col].head(3).tolist()}")
                    df[col] = df[actual_col].apply(to_hours)
                    print(f"DEBUG: After conversion - Sample values: {df[col].head(3).tolist()}")
                    break
    
    # Normalize percentage columns
    percentage_columns = ["Attended % to Date.", "% Missed"]
    
    for col in percentage_columns:
        if col in df.columns:
            print(f"DEBUG: Normalizing percentage column '{col}' - Sample values before: {df[col].head(3).tolist()}")
            df[col] = df[col].apply(normalize_pct)
            print(f"DEBUG: After normalization - Sample values: {df[col].head(3).tolist()}")
        else:
            # Try case-insensitive search
            for actual_col in df.columns:
                if col.lower().strip() == actual_col.lower().strip():
                    print(f"DEBUG: Found column '{actual_col}' (case variation of '{col}')")
                    print(f"DEBUG: Sample values before: {df[actual_col].head(3).tolist()}")
                    df[col] = df[actual_col].apply(normalize_pct)
                    print(f"DEBUG: After normalization - Sample values: {df[col].head(3).tolist()}")
                    break
    
    # Replace invalid or missing values (but preserve valid zeros)
    # Only replace NaN, Infinity, and -Infinity
    df = df.replace([np.inf, -np.inf], 0)
    # Fill NaN values with 0 only for numeric columns
    numeric_cols = df.select_dtypes(include=[np.number]).columns
    for col in numeric_cols:
        df[col] = df[col].fillna(0)
    
    print(f"=== END DEBUG: normalize_attendance_data ===\n")
    
    return df


def normalize_percentage(value: float, max_value: float = 1.0) -> float:
    """
    Normalize percentage to 0-100 range.
    
    If max_value <= 1.0, multiply by 100.
    """
    if pd.isna(value):
        return 0.0
    
    if max_value <= 1.0:
        return float(value) * 100.0
    return float(value)


def extract_hyperlink_from_cell(cell: Cell) -> Optional[str]:
    """Extract hyperlink from an openpyxl cell if present."""
    if cell.hyperlink and cell.hyperlink.target:
        return cell.hyperlink.target
    return None


def load_refined_data(file_bytes: bytes) -> Tuple[pd.DataFrame, pd.DataFrame, Dict[str, str]]:
    """
    Simplified data loading - only reads refined columns from Excel sheets.
    
    Grades Sheet: Student#, Student Name, Program Name, Program Grade
    Attendance Sheet: Student#, Student Name, Attended % to Date.
    
    Args:
        file_bytes: Raw bytes of the Excel file
    
    Returns:
        Tuple of (grades_df, attendance_df, name_hyperlinks_dict)
    """
    from io import BytesIO
    from openpyxl import load_workbook
    import pandas as pd
    
    workbook = load_workbook(filename=BytesIO(file_bytes), data_only=False)
    excel_file = BytesIO(file_bytes)
    
    # Find sheet names
    grades_sheet_name = None
    attendance_sheet_name = None
    
    for sheet_name in workbook.sheetnames:
        if 'grade' in sheet_name.lower():
            grades_sheet_name = sheet_name
        elif 'attend' in sheet_name.lower():
            attendance_sheet_name = sheet_name
    
    if not grades_sheet_name:
        raise ValueError(f"Could not find grades sheet. Available sheets: {workbook.sheetnames}")
    if not attendance_sheet_name:
        raise ValueError(f"Could not find attendance sheet. Available sheets: {workbook.sheetnames}")
    
    # Load DataFrames
    grades_df = pd.read_excel(excel_file, sheet_name=grades_sheet_name, engine='openpyxl')
    excel_file.seek(0)  # Reset file pointer
    attendance_df = pd.read_excel(excel_file, sheet_name=attendance_sheet_name, engine='openpyxl')
    
    # Extract hyperlinks from Student Name columns
    name_hyperlinks = {}
    
    # Extract from grades sheet
    if grades_sheet_name in workbook.sheetnames:
        grades_sheet = workbook[grades_sheet_name]
        # Find column indices
        header_row = None
        student_id_col = None
        student_name_col = None
        
        for row_idx, row in enumerate(grades_sheet.iter_rows(min_row=1, max_row=20), start=1):
            for col_idx, cell in enumerate(row, start=1):
                cell_value = str(cell.value).strip() if cell.value else ''
                if 'Student#' in cell_value or 'Student #' in cell_value:
                    header_row = row_idx
                    student_id_col = col_idx
                elif 'Student Name' in cell_value:
                    student_name_col = col_idx
            
            if header_row and student_id_col and student_name_col:
                break
        
        # Extract hyperlinks
        if header_row and student_id_col and student_name_col:
            for row in grades_sheet.iter_rows(min_row=header_row + 1):
                if len(row) < max(student_id_col, student_name_col):
                    continue
                try:
                    student_id_cell = row[student_id_col - 1]
                    student_name_cell = row[student_name_col - 1]
                    
                    if student_id_cell.value and student_name_cell.value:
                        try:
                            student_id = str(int(float(student_id_cell.value)))
                        except (ValueError, TypeError):
                            student_id = str(student_id_cell.value).strip()
                        hyperlink = extract_hyperlink_from_cell(student_name_cell)
                        if hyperlink:
                            name_hyperlinks[student_id] = hyperlink
                except (IndexError, TypeError):
                    continue
    
    # Normalize and rename columns in grades_df
    grades_df = normalize_and_rename_columns(grades_df, 'grades')
    
    # Keep only required columns: Student#, Student Name, Program Name, Program Grade
    required_grades_cols = []
    for col in ['Student#', 'Student Name', 'Program Name', 'current overall Program Grade']:
        if col in grades_df.columns:
            required_grades_cols.append(col)
        else:
            # Try to find similar column
            for actual_col in grades_df.columns:
                if normalize_col_name(actual_col) == normalize_col_name(col):
                    required_grades_cols.append(actual_col)
                    break
    
    if len(required_grades_cols) < 4:
        raise ValueError(f"Missing required columns in grades sheet. Found: {list(grades_df.columns)}, Required: Student#, Student Name, Program Name, Program Grade")
    
    # Select and rename columns
    grades_df = grades_df[required_grades_cols].copy()
    grades_df.columns = ['Student#', 'Student Name', 'Program Name', 'Program Grade']
    
    # Normalize and rename columns in attendance_df
    attendance_df = normalize_and_rename_columns(attendance_df, 'attendance')
    
    # Keep only required columns: Student#, Student Name, Attended % to Date.
    required_attendance_cols = []
    for col in ['Student#', 'Student Name', 'Attended % to Date.']:
        if col in attendance_df.columns:
            required_attendance_cols.append(col)
        else:
            # Try to find similar column
            for actual_col in attendance_df.columns:
                if normalize_col_name(actual_col) == normalize_col_name(col):
                    required_attendance_cols.append(actual_col)
                    break
    
    if len(required_attendance_cols) < 3:
        raise ValueError(f"Missing required columns in attendance sheet. Found: {list(attendance_df.columns)}, Required: Student#, Student Name, Attended % to Date.")
    
    # Select and rename columns
    attendance_df = attendance_df[required_attendance_cols].copy()
    attendance_df.columns = ['Student#', 'Student Name', 'Attended % to Date.']
    
    # Clean and standardize formatting
    # Remove summary rows
    grades_df = grades_df[~grades_df['Student Name'].astype(str).str.contains('Total|Summary', case=False, na=False)]
    attendance_df = attendance_df[~attendance_df['Student Name'].astype(str).str.contains('Total|Summary', case=False, na=False)]
    
    # Remove rows with missing Student# or Student Name
    grades_df = grades_df.dropna(subset=['Student#', 'Student Name'])
    attendance_df = attendance_df.dropna(subset=['Student#', 'Student Name'])
    
    # Normalize Program Grade (remove %, convert to float 0-100)
    grades_df['Program Grade'] = grades_df['Program Grade'].apply(normalize_pct)
    
    # Normalize Attended % to Date. (remove %, convert to float 0-100)
    attendance_df['Attended % to Date.'] = attendance_df['Attended % to Date.'].apply(normalize_pct)
    
    # Standardize Student# to string (remove .0, strip whitespace)
    grades_df['Student#'] = grades_df['Student#'].astype(str).str.strip().str.replace('.0', '', regex=False)
    attendance_df['Student#'] = attendance_df['Student#'].astype(str).str.strip().str.replace('.0', '', regex=False)
    
    # Standardize Student Name
    grades_df['Student Name'] = grades_df['Student Name'].astype(str).str.strip()
    attendance_df['Student Name'] = attendance_df['Student Name'].astype(str).str.strip()
    
    # Rename to standardized column names
    grades_df = grades_df.rename(columns={
        'Student#': 'Student ID',
        'Program Name': 'Program',
        'Program Grade': 'Grade %'
    })
    
    attendance_df = attendance_df.rename(columns={
        'Student#': 'Student ID',
        'Attended % to Date.': 'Attendance %'
    })
    
    print(f"DEBUG: Simplified loading - Grades: {len(grades_df)} rows, Attendance: {len(attendance_df)} rows")
    print(f"DEBUG: Grades columns: {list(grades_df.columns)}")
    print(f"DEBUG: Attendance columns: {list(attendance_df.columns)}")
    
    return grades_df, attendance_df, name_hyperlinks


def prepare_combined_dataset(grades_df: pd.DataFrame, attendance_df: pd.DataFrame) -> pd.DataFrame:
    """
    Merge grades and attendance data using outer join.
    
    Args:
        grades_df: DataFrame with Student ID, Student Name, Program, Grade %
        attendance_df: DataFrame with Student ID, Student Name, Attendance %
    
    Returns:
        Merged DataFrame with all students
    """
    # Outer join to include students missing in one sheet
    merged_df = pd.merge(
        grades_df,
        attendance_df,
        on=['Student ID', 'Student Name'],
        how='outer',
        suffixes=('_grades', '_attendance')
    )
    
    # Handle merged columns - prioritize non-null values
    if 'Grade %_grades' in merged_df.columns and 'Grade %_attendance' in merged_df.columns:
        merged_df['Grade %'] = merged_df['Grade %_grades'].fillna(merged_df['Grade %_attendance'])
        merged_df = merged_df.drop(columns=['Grade %_grades', 'Grade %_attendance'])
    elif 'Grade %_grades' in merged_df.columns:
        merged_df['Grade %'] = merged_df['Grade %_grades']
        merged_df = merged_df.drop(columns=['Grade %_grades'])
    elif 'Grade %_attendance' in merged_df.columns:
        merged_df['Grade %'] = merged_df['Grade %_attendance']
        merged_df = merged_df.drop(columns=['Grade %_attendance'])
    
    if 'Attendance %_grades' in merged_df.columns and 'Attendance %_attendance' in merged_df.columns:
        merged_df['Attendance %'] = merged_df['Attendance %_attendance'].fillna(merged_df['Attendance %_grades'])
        merged_df = merged_df.drop(columns=['Attendance %_grades', 'Attendance %_attendance'])
    elif 'Attendance %_grades' in merged_df.columns:
        merged_df['Attendance %'] = merged_df['Attendance %_grades']
        merged_df = merged_df.drop(columns=['Attendance %_grades'])
    elif 'Attendance %_attendance' in merged_df.columns:
        merged_df['Attendance %'] = merged_df['Attendance %_attendance']
        merged_df = merged_df.drop(columns=['Attendance %_attendance'])
    
    if 'Program_grades' in merged_df.columns and 'Program_attendance' in merged_df.columns:
        merged_df['Program'] = merged_df['Program_grades'].fillna(merged_df['Program_attendance'])
        merged_df = merged_df.drop(columns=['Program_grades', 'Program_attendance'])
    elif 'Program_grades' in merged_df.columns:
        merged_df['Program'] = merged_df['Program_grades']
        merged_df = merged_df.drop(columns=['Program_grades'])
    elif 'Program_attendance' in merged_df.columns:
        merged_df['Program'] = merged_df['Program_attendance']
        merged_df = merged_df.drop(columns=['Program_attendance'])
    
    # Fill missing values gracefully
    merged_df['Grade %'] = merged_df['Grade %'].fillna(0.0)
    merged_df['Attendance %'] = merged_df['Attendance %'].fillna(0.0)
    merged_df['Program'] = merged_df['Program'].fillna('Unknown')
    
    # Determine data status
    merged_df['_has_grade'] = merged_df['Grade %'] > 0
    merged_df['_has_attendance'] = merged_df['Attendance %'] > 0
    
    def determine_status(row):
        if row['_has_grade'] and row['_has_attendance']:
            return 'Complete'
        elif row['_has_grade'] and not row['_has_attendance']:
            return 'Missing Attendance'
        elif not row['_has_grade'] and row['_has_attendance']:
            return 'Missing Grade'
        else:
            return 'Missing Both'
    
    merged_df['data_status'] = merged_df.apply(determine_status, axis=1)
    merged_df = merged_df.drop(columns=['_has_grade', '_has_attendance'])
    
    # Ensure Student ID and Student Name are strings
    merged_df['Student ID'] = merged_df['Student ID'].astype(str)
    merged_df['Student Name'] = merged_df['Student Name'].astype(str)
    
    print(f"DEBUG: Merged dataset: {len(merged_df)} rows")
    print(f"DEBUG: Merged columns: {list(merged_df.columns)}")
    
    return merged_df


def load_excel(file_bytes: bytes) -> Tuple[pd.DataFrame, pd.DataFrame, Dict[str, str]]:
    """
    Load Excel file and extract data from both sheets.
    
    Expected Excel structure:
    - Sheet 1: "Students Grade"
      - Student#: numeric (unique student ID)
      - Student Name: string
      - Program Name: string
      - current overall Program Grade: float (0-1 decimal or 0-100 percent)
    
    - Sheet 2: "Students attendance " (note trailing space)
      - Student#: numeric (same ID key as Grades sheet)
      - Student Name: string (some have hyperlinks to Campus Login)
      - Scheduled Hours to Date: string "HH:MM" (e.g., "90:00")
      - Attended Hours to Date: string "HH:MM" (e.g., "89:45")
      - Attended % to Date.: float 0-1 (e.g., 0.997222 = 99.7%)
      - Missed Hours to Date: string "HH:MM" (e.g., "5:00")
      - % Missed: float 0-1
      - Missed Minus Excused to date: string "HH:MM" or number (e.g., "0:15" or 0)
    
    Args:
        file_bytes: Raw bytes of the Excel file
    
    Returns:
        Tuple of (grades_df, attendance_df, name_hyperlinks_dict)
        where name_hyperlinks_dict maps student_id -> hyperlink_url
    """
    # Load workbook to extract hyperlinks
    workbook = load_workbook(filename=BytesIO(file_bytes), data_only=False)
    
    # Extract hyperlinks from Grades sheet
    name_hyperlinks = {}
    
    if 'Students Grade' in workbook.sheetnames:
        grades_sheet = workbook['Students Grade']
        
        # Find Student# and Student Name columns
        header_row = None
        student_id_col = None
        student_name_col = None
        
        for row_idx, row in enumerate(grades_sheet.iter_rows(min_row=1, max_row=20), start=1):
            for col_idx, cell in enumerate(row, start=1):
                cell_value = str(cell.value).strip() if cell.value else ''
                if 'Student#' in cell_value or 'Student #' in cell_value:
                    header_row = row_idx
                    student_id_col = col_idx
                elif 'Student Name' in cell_value:
                    student_name_col = col_idx
            
            if header_row and student_id_col and student_name_col:
                break
        
        # Extract hyperlinks
        if header_row and student_id_col and student_name_col:
            for row in grades_sheet.iter_rows(min_row=header_row + 1):
                # Check if row has enough columns
                if len(row) < max(student_id_col, student_name_col):
                    continue
                try:
                    student_id_cell = row[student_id_col - 1]
                    student_name_cell = row[student_name_col - 1]
                    
                    if student_id_cell.value and student_name_cell.value:
                        # Convert Student# to numeric for consistent matching
                        try:
                            student_id = str(int(float(student_id_cell.value)))
                        except (ValueError, TypeError):
                            student_id = str(student_id_cell.value).strip()
                        hyperlink = extract_hyperlink_from_cell(student_name_cell)
                        if hyperlink:
                            name_hyperlinks[student_id] = hyperlink
                except (IndexError, TypeError) as e:
                    # Skip rows that don't have enough columns or have invalid data
                    continue
    
    # Load dataframes using pandas
    excel_file = BytesIO(file_bytes)
    
    # Load Grades sheet
    try:
        grades_df = pd.read_excel(excel_file, sheet_name='Students Grade', engine='openpyxl')
    except ValueError as e:
        raise ValueError(f"Could not find 'Students Grade' sheet. Available sheets: {workbook.sheetnames}")
    
    # DEBUG: Print raw Excel data to understand structure
    print(f"\n=== DEBUG: Raw Grades DataFrame from Excel ===")
    print(f"Shape: {grades_df.shape}")
    print(f"Columns: {list(grades_df.columns)}")
    if len(grades_df) > 0:
        print(f"First 3 rows (all columns):")
        for idx in range(min(3, len(grades_df))):
            print(f"  Row {idx}:")
            for col in grades_df.columns:
                val = grades_df.iloc[idx][col]
                print(f"    {col}: {val} (type: {type(val).__name__})")
    print(f"=== END DEBUG: Raw Grades DataFrame ===\n")
    
    # Reset file pointer for second sheet
    excel_file.seek(0)
    
    # Load Attendance sheet (note the trailing space) and extract hyperlinks
    attendance_sheet_name = None
    for sheet_name in workbook.sheetnames:
        if 'attendance' in sheet_name.lower():
            attendance_sheet_name = sheet_name
            break
    
    if not attendance_sheet_name:
        raise ValueError(f"Could not find attendance sheet. Available sheets: {workbook.sheetnames}")
    
    # Extract hyperlinks from Attendance sheet
    attendance_sheet = workbook[attendance_sheet_name]
    attendance_hyperlinks = {}
    
    # Find Student# and Student Name columns in attendance sheet
    att_header_row = None
    att_student_id_col = None
    att_student_name_col = None
    
    for row_idx, row in enumerate(attendance_sheet.iter_rows(min_row=1, max_row=20), start=1):
        for col_idx, cell in enumerate(row, start=1):
            cell_value = str(cell.value).strip() if cell.value else ''
            if 'Student#' in cell_value or 'Student #' in cell_value:
                att_header_row = row_idx
                att_student_id_col = col_idx
            elif 'Student Name' in cell_value:
                att_student_name_col = col_idx
        
        if att_header_row and att_student_id_col and att_student_name_col:
            break
    
    # Extract hyperlinks from attendance sheet
    if att_header_row and att_student_id_col and att_student_name_col:
        for row in attendance_sheet.iter_rows(min_row=att_header_row + 1):
            # Check if row has enough columns
            if len(row) < max(att_student_id_col, att_student_name_col):
                continue
            try:
                student_id_cell = row[att_student_id_col - 1]
                student_name_cell = row[att_student_name_col - 1]
                
                if student_id_cell.value and student_name_cell.value:
                    # Convert Student# to numeric for consistent matching
                    try:
                        student_id = str(int(float(student_id_cell.value)))
                    except (ValueError, TypeError):
                        student_id = str(student_id_cell.value).strip()
                    hyperlink = extract_hyperlink_from_cell(student_name_cell)
                    if hyperlink:
                        attendance_hyperlinks[student_id] = hyperlink
                        # Attendance sheet hyperlinks take precedence over grades sheet
                        name_hyperlinks[student_id] = hyperlink
            except IndexError:
                # Skip rows that don't have enough columns
                continue
    
    # Load attendance dataframe
    attendance_df = pd.read_excel(excel_file, sheet_name=attendance_sheet_name, engine='openpyxl')
    
    # DEBUG: Print raw Excel data to understand structure
    print(f"\n=== DEBUG: Raw Attendance DataFrame from Excel ===")
    print(f"Shape: {attendance_df.shape}")
    print(f"Columns: {list(attendance_df.columns)}")
    if len(attendance_df) > 0:
        print(f"First 3 rows (all columns):")
        for idx in range(min(3, len(attendance_df))):
            print(f"  Row {idx}:")
            for col in attendance_df.columns:
                val = attendance_df.iloc[idx][col]
                print(f"    {col}: {val} (type: {type(val).__name__})")
    print(f"=== END DEBUG: Raw Attendance DataFrame ===\n")
    
    # Normalize and rename columns to handle variations BEFORE validation
    print(f"DEBUG: Original grades columns: {list(grades_df.columns)}")
    grades_df = normalize_and_rename_columns(grades_df, "grades")
    print(f"DEBUG: Normalized grades columns: {list(grades_df.columns)}")
    
    # DEBUG: After normalization, check what's in Student Name and Student# columns
    if 'Student Name' in grades_df.columns and 'Student#' in grades_df.columns:
        print(f"\n=== DEBUG: After normalization - Grades DataFrame ===")
        print(f"Student# sample (first 5): {grades_df['Student#'].head(5).tolist()}")
        print(f"Student Name sample (first 5): {grades_df['Student Name'].head(5).tolist()}")
        # Check if Student Name contains numeric values (misalignment indicator)
        student_name_sample = grades_df['Student Name'].head(10)
        numeric_names = student_name_sample.astype(str).str.match(r'^\d+$', na=False).sum()
        if numeric_names > 0:
            print(f"WARNING: {numeric_names} out of 10 Student Names appear to be numeric (IDs)!")
            print(f"  Sample numeric 'names': {student_name_sample[student_name_sample.astype(str).str.match(r'^\d+$', na=False)].head(3).tolist()}")
        print(f"=== END DEBUG ===\n")
    
    print(f"DEBUG: Original attendance columns: {list(attendance_df.columns)}")
    attendance_df = normalize_and_rename_columns(attendance_df, "attendance")
    print(f"DEBUG: Normalized attendance columns: {list(attendance_df.columns)}")
    
    # DEBUG: After normalization, check what's in Student Name and Student# columns
    if 'Student Name' in attendance_df.columns and 'Student#' in attendance_df.columns:
        print(f"\n=== DEBUG: After normalization - Attendance DataFrame ===")
        print(f"Student# sample (first 5): {attendance_df['Student#'].head(5).tolist()}")
        print(f"Student Name sample (first 5): {attendance_df['Student Name'].head(5).tolist()}")
        # Check if Student Name contains numeric values (misalignment indicator)
        student_name_sample = attendance_df['Student Name'].head(10)
        numeric_names = student_name_sample.astype(str).str.match(r'^\d+$', na=False).sum()
        if numeric_names > 0:
            print(f"WARNING: {numeric_names} out of 10 Student Names appear to be numeric (IDs)!")
            print(f"  Sample numeric 'names': {attendance_df['Student Name'][attendance_df['Student Name'].astype(str).str.match(r'^\d+$', na=False)].head(3).tolist()}")
        print(f"=== END DEBUG ===\n")
    
    # Also strip whitespace from column names (additional cleanup)
    # Ensure columns is an Index before using .str
    try:
        if hasattr(grades_df.columns, 'str'):
            grades_df.columns = grades_df.columns.str.strip()
        else:
            # Fallback: manually strip column names
            grades_df.columns = [str(col).strip() for col in grades_df.columns]
    except Exception as e:
        print(f"WARNING: Could not strip grades_df column names: {e}")
        grades_df.columns = [str(col).strip() for col in grades_df.columns]
    
    try:
        if hasattr(attendance_df.columns, 'str'):
            attendance_df.columns = attendance_df.columns.str.strip()
        else:
            # Fallback: manually strip column names
            attendance_df.columns = [str(col).strip() for col in attendance_df.columns]
    except Exception as e:
        print(f"WARNING: Could not strip attendance_df column names: {e}")
        attendance_df.columns = [str(col).strip() for col in attendance_df.columns]
    
    # Debug: Print raw attendance DataFrame info
    print(f"\n=== DEBUG: Raw attendance DataFrame after pd.read_excel ===")
    print(f"Shape: {attendance_df.shape}")
    print(f"Columns: {list(attendance_df.columns)}")
    if 'Attended % to Date.' in attendance_df.columns:
        print(f"'Attended % to Date.' dtype: {attendance_df['Attended % to Date.'].dtype}")
        print(f"'Attended % to Date.' sample values (raw): {attendance_df['Attended % to Date.'].head(5).tolist()}")
        print(f"'Attended % to Date.' non-null count: {attendance_df['Attended % to Date.'].notna().sum()}")
        print(f"'Attended % to Date.' null count: {attendance_df['Attended % to Date.'].isna().sum()}")
    else:
        print("WARNING: 'Attended % to Date.' column NOT found in raw DataFrame!")
        # Try to find similar column names
        for col in attendance_df.columns:
            if 'attended' in col.lower() and '%' in col:
                print(f"Found similar column: '{col}' with values: {attendance_df[col].head(3).tolist()}")
    print(f"=== END DEBUG: Raw attendance DataFrame ===\n")
    
    # Clean Student# in both DataFrames to ensure matching
    print("=== DEBUG: Cleaning Student# columns ===")
    grades_df = clean_student_id(grades_df)
    attendance_df = clean_student_id(attendance_df)
    print(f"After cleaning - grades_df Student# sample: {grades_df['Student#'].head(3).tolist() if 'Student#' in grades_df.columns else 'N/A'}")
    print(f"After cleaning - attendance_df Student# sample: {attendance_df['Student#'].head(3).tolist() if 'Student#' in attendance_df.columns else 'N/A'}")
    
    # Clean attendance DataFrame (remove totals, convert times, normalize percentages)
    print("=== DEBUG: Cleaning attendance DataFrame ===")
    attendance_df = clean_attendance_df(attendance_df)
    print(f"After cleaning attendance_df, shape: {attendance_df.shape}")
    
    # Drop rows missing Student# or Student Name
    print("=== DEBUG: Dropping rows with missing Student# or Student Name ===")
    initial_grades = len(grades_df)
    initial_attendance = len(attendance_df)
    
    # Ensure Student Name exists in both DataFrames before dropping
    if 'Student Name' not in grades_df.columns:
        print("ERROR: 'Student Name' column missing in grades_df after normalization!")
        print(f"Available columns in grades_df: {list(grades_df.columns)}")
        # Try to create it
        if len(grades_df) > 0:
            grades_df['Student Name'] = 'Unknown'
            print("WARNING: Created placeholder 'Student Name' column in grades_df")
        else:
            raise ValueError(f"'Student Name' column missing in grades DataFrame. Available columns: {list(grades_df.columns)}")
    
    if 'Student Name' not in attendance_df.columns:
        print("ERROR: 'Student Name' column missing in attendance_df after normalization!")
        print(f"Available columns in attendance_df: {list(attendance_df.columns)}")
        # Try to create it
        if len(attendance_df) > 0:
            attendance_df['Student Name'] = 'Unknown'
            print("WARNING: Created placeholder 'Student Name' column in attendance_df")
        else:
            raise ValueError(f"'Student Name' column missing in attendance DataFrame. Available columns: {list(attendance_df.columns)}")
    
    # Now safely drop rows with missing values
    drop_cols_grades = []
    drop_cols_attendance = []
    
    if 'Student#' in grades_df.columns:
        drop_cols_grades.append('Student#')
    if 'Student Name' in grades_df.columns:
        drop_cols_grades.append('Student Name')
    
    if 'Student#' in attendance_df.columns:
        drop_cols_attendance.append('Student#')
    if 'Student Name' in attendance_df.columns:
        drop_cols_attendance.append('Student Name')
    
    if drop_cols_grades:
        grades_df = grades_df.dropna(subset=drop_cols_grades)
        print(f"Grades: Dropped {initial_grades - len(grades_df)} rows with missing {drop_cols_grades}")
    
    if drop_cols_attendance:
        attendance_df = attendance_df.dropna(subset=drop_cols_attendance)
        print(f"Attendance: Dropped {initial_attendance - len(attendance_df)} rows with missing {drop_cols_attendance}")
    
    # Ensure both DataFrames use the same data type for Student#
    if 'Student#' in grades_df.columns and 'Student#' in attendance_df.columns:
        try:
            grades_df["Student#"] = grades_df["Student#"].fillna(0).astype(int)
            attendance_df["Student#"] = attendance_df["Student#"].fillna(0).astype(int)
            print(f"Converted Student# to int - grades sample: {grades_df['Student#'].head(3).tolist()}, attendance sample: {attendance_df['Student#'].head(3).tolist()}")
        except (ValueError, TypeError) as e:
            print(f"Warning: Could not convert Student# to int: {e}, keeping as numeric")
            grades_df["Student#"] = pd.to_numeric(grades_df["Student#"], errors='coerce').fillna(0)
            attendance_df["Student#"] = pd.to_numeric(attendance_df["Student#"], errors='coerce').fillna(0)
    
    # Add Campus Login URL column to attendance_df (after column normalization)
    # First, ensure Student# is numeric for consistent matching
    if 'Student#' in attendance_df.columns:
        # Convert to numeric, then to string for mapping
        try:
            student_id_series = safe_get_series(attendance_df, 'Student#')
            student_ids = pd.to_numeric(student_id_series, errors='coerce').fillna(0).astype(int).astype(str)
        except (ValueError, TypeError):
            student_id_series = safe_get_series(attendance_df, 'Student#')
            student_ids = student_id_series.astype(str).str.strip()
        attendance_df['Campus Login URL'] = student_ids.map(
            lambda x: attendance_hyperlinks.get(x, None)
        )
    else:
        # Find Student# column
        student_id_col = None
        for col in attendance_df.columns:
            if 'student#' in col.lower().strip():
                student_id_col = col
                break
        if student_id_col:
            try:
                student_id_series = safe_get_series(attendance_df, student_id_col)
                student_ids = pd.to_numeric(student_id_series, errors='coerce').fillna(0).astype(int).astype(str)
            except (ValueError, TypeError):
                student_id_series = safe_get_series(attendance_df, student_id_col)
                student_ids = student_id_series.astype(str).str.strip()
            attendance_df['Campus Login URL'] = student_ids.map(
                lambda x: attendance_hyperlinks.get(x, None)
            )
        else:
            attendance_df['Campus Login URL'] = None
    
    # Validate required columns - only check for the columns we actually need
    # Grades: Student#, Student Name, Program Name, current overall Program Grade
    # Attendance: Student#, Student Name, Attended % to Date.
    required_grades_cols = ['Student Name', 'Program Name', 'current overall Program Grade']
    required_attendance_cols = ['Student Name', 'Attended % to Date.']
    
    # Check for column name variations (case-insensitive, handle whitespace, dots, %)
    # Normalize column names for comparison
    def normalize_col_name_for_validation(col_name):
        return re.sub(r'\s+', ' ', str(col_name).lower().strip().replace('.', '').replace('%', '').replace(',', ''))
    
    grades_cols_normalized = {normalize_col_name_for_validation(col): col for col in grades_df.columns}
    attendance_cols_normalized = {normalize_col_name_for_validation(col): col for col in attendance_df.columns}
    
    missing_grades = []
    missing_attendance = []
    
    for req_col in required_grades_cols:
        req_col_normalized = normalize_col_name_for_validation(req_col)
        if req_col_normalized not in grades_cols_normalized:
            missing_grades.append(req_col)
    
    for req_col in required_attendance_cols:
        req_col_normalized = normalize_col_name_for_validation(req_col)
        if req_col_normalized not in attendance_cols_normalized:
            missing_attendance.append(req_col)
    
    if missing_grades or missing_attendance:
        error_msg = "Missing required columns:\n"
        if missing_grades:
            error_msg += f"  Grades sheet: {missing_grades}\n"
        if missing_attendance:
            error_msg += f"  Attendance sheet: {missing_attendance}\n"
        error_msg += f"\nFound columns:\n"
        error_msg += f"  Grades: {list(grades_df.columns)}\n"
        error_msg += f"  Attendance: {list(attendance_df.columns)}\n"
        error_msg += f"\nNote: Student# is optional and will be auto-generated if missing."
        raise ValueError(error_msg)
    
    return grades_df, attendance_df, name_hyperlinks


def normalize_data(grades_df: pd.DataFrame, attendance_df: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    Normalize data: convert percentages, durations, and clean student IDs.
    
    Args:
        grades_df: Raw grades dataframe
        attendance_df: Raw attendance dataframe
    
    Returns:
        Tuple of normalized dataframes
    """
    # Normalize Grades sheet
    grades_normalized = grades_df.copy()
    
    # CRITICAL: Ensure Student Name is preserved before any processing
    if 'Student Name' not in grades_normalized.columns:
        print(f"ERROR: 'Student Name' missing in grades_df before normalization!")
        print(f"Available columns: {list(grades_normalized.columns)}")
        raise ValueError(f"'Student Name' column missing in grades DataFrame. Available columns: {list(grades_normalized.columns)}")
    
    print(f"DEBUG: normalize_data - grades_df has Student Name: {grades_normalized['Student Name'].head(3).tolist() if len(grades_normalized) > 0 else 'EMPTY'}")

    # Find Student# column (case-insensitive)
    student_id_col = None
    for col in grades_normalized.columns:
        if 'student#' in col.lower().strip():
            student_id_col = col
            break

    if student_id_col:
        # Convert Student# to numeric (int) for consistent matching
        try:
            # First try to convert to numeric, handling any string values
            student_ids = pd.to_numeric(grades_normalized[student_id_col], errors='coerce')
            # Replace NaN with 0, but keep track of which ones failed
            grades_normalized['Student#'] = student_ids.fillna(0).astype(int)
            # If we got all zeros, the conversion probably failed - use string instead
            student_id_series = safe_get_series(grades_normalized, 'Student#')
            if (student_id_series == 0).all():
                print("WARNING: All Student# values converted to 0 in grades, using string format instead")
                source_series = safe_get_series(grades_normalized, student_id_col)
                grades_normalized['Student#'] = source_series.astype(str).str.strip()
            else:
                print(f"Successfully converted grades Student# to int, sample: {student_id_series.head(3).tolist()}")
        except (ValueError, TypeError) as e:
            print(f"Warning: Could not convert grades Student# to numeric: {e}, using string format")
            # Fallback to string if conversion fails
            source_series = safe_get_series(grades_normalized, student_id_col)
            grades_normalized['Student#'] = source_series.astype(str).str.strip()
    
    # Normalize grade percentage (find column case-insensitively)
    # Try multiple possible column names: "current overall Program Grade", "Program Grade", etc.
    grade_col = None
    possible_grade_cols = [
        'current overall program grade',
        'program grade',
        'grade',
        'overall grade',
        'final grade'
    ]
    
    for possible_name in possible_grade_cols:
        for col in grades_normalized.columns:
            if possible_name in col.lower().strip():
                grade_col = col
                break
        if grade_col:
            break
    
    if grade_col:
        # Apply normalize_pct to convert 0-1 to 0-100, or handle percentage strings
        print(f"DEBUG: Normalizing grade column '{grade_col}' - Sample values before: {grades_normalized[grade_col].head(3).tolist()}")
        # Handle percentage strings (e.g., "88%") and normalize
        grades_normalized['grade_pct'] = grades_normalized[grade_col].apply(normalize_pct)
        print(f"DEBUG: After normalization - Sample values: {grades_normalized['grade_pct'].head(3).tolist()}")
    else:
        print(f"WARNING: Grade column not found. Available columns: {list(grades_normalized.columns)}")
        grades_normalized['grade_pct'] = 0.0
    
    # Normalize Attendance sheet - only process required columns
    attendance_normalized = attendance_df.copy()
    
    # CRITICAL: Ensure Student Name is preserved before any processing
    if 'Student Name' not in attendance_normalized.columns:
        print(f"ERROR: 'Student Name' missing in attendance_df before normalization!")
        print(f"Available columns: {list(attendance_normalized.columns)}")
        raise ValueError(f"'Student Name' column missing in attendance DataFrame. Available columns: {list(attendance_normalized.columns)}")
    
    print(f"DEBUG: normalize_data - attendance_df has Student Name: {attendance_normalized['Student Name'].head(3).tolist() if len(attendance_normalized) > 0 else 'EMPTY'}")

    # Preserve Campus Login URL column if it exists
    campus_login_url_col = None
    for col in attendance_normalized.columns:
        if 'campus login url' in col.lower().strip():
            campus_login_url_col = col
            break

    # Find Student# column (case-insensitive)
    student_id_col = None
    for col in attendance_normalized.columns:
        if 'student#' in col.lower().strip():
            student_id_col = col
            break

    if student_id_col:
        # Convert Student# to numeric (int) for consistent matching
        try:
            # First try to convert to numeric, handling any string values
            student_ids = pd.to_numeric(attendance_normalized[student_id_col], errors='coerce')
            # Replace NaN with 0, but keep track of which ones failed
            attendance_normalized['Student#'] = student_ids.fillna(0).astype(int)
            # If we got all zeros, the conversion probably failed - use string instead
            student_id_series = safe_get_series(attendance_normalized, 'Student#')
            if (student_id_series == 0).all():
                print("WARNING: All Student# values converted to 0 in attendance, using string format instead")
                source_series = safe_get_series(attendance_normalized, student_id_col)
                attendance_normalized['Student#'] = source_series.astype(str).str.strip()
            else:
                print(f"Successfully converted attendance Student# to int, sample: {student_id_series.head(3).tolist()}")
        except (ValueError, TypeError) as e:
            print(f"Warning: Could not convert attendance Student# to numeric: {e}, using string format")
            # Fallback to string if conversion fails
            source_series = safe_get_series(attendance_normalized, student_id_col)
            attendance_normalized['Student#'] = source_series.astype(str).str.strip()

    # Preserve Campus Login URL column
    if campus_login_url_col and campus_login_url_col != 'Campus Login URL':
        attendance_normalized['Campus Login URL'] = attendance_normalized[campus_login_url_col]
    elif 'Campus Login URL' not in attendance_normalized.columns:
        attendance_normalized['Campus Login URL'] = None

    # PREPROCESSING STEP: Automatically clean and normalize attendance data
    # Note: clean_attendance_df was already called in load_excel, but we may need to re-normalize
    # if the data was modified. Check if columns still need conversion.
    print(f"\n=== DEBUG: Before normalize_attendance_data ===")
    print(f"Columns in attendance_normalized: {list(attendance_normalized.columns)}")
    if 'Attended % to Date.' in attendance_normalized.columns:
        print(f"Sample 'Attended % to Date.' values: {attendance_normalized['Attended % to Date.'].head(3).tolist()}")
    
    # Only normalize if columns haven't been normalized yet (check if values are still 0-1 range)
    needs_normalization = False
    if 'Attended % to Date.' in attendance_normalized.columns:
        sample_values = attendance_normalized['Attended % to Date.'].dropna()
        if len(sample_values) > 0:
            max_val = sample_values.max()
            if max_val <= 1.0 and max_val > 0:
                needs_normalization = True
                print(f"DEBUG: Values still in 0-1 range (max: {max_val}), will normalize")
    
    if needs_normalization:
        attendance_normalized = normalize_attendance_data(attendance_normalized)
    else:
        print("DEBUG: Attendance data already normalized, skipping normalize_attendance_data")
    
    # Extract attendance_pct from normalized "Attended % to Date." column
    if 'Attended % to Date.' in attendance_normalized.columns:
        attendance_normalized['attendance_pct'] = attendance_normalized['Attended % to Date.']
        print(f"DEBUG: Set attendance_pct from 'Attended % to Date.' - Sample: {attendance_normalized['attendance_pct'].head(3).tolist()}")
    else:
        # Try case-insensitive search for the column
        found_col = None
        for col in attendance_normalized.columns:
            if 'attended' in col.lower() and '%' in col and 'date' in col.lower():
                found_col = col
                print(f"DEBUG: Found attendance percentage column: '{found_col}'")
                attendance_normalized['attendance_pct'] = attendance_normalized[found_col]
                print(f"DEBUG: Set attendance_pct from '{found_col}' - Sample: {attendance_normalized['attendance_pct'].head(3).tolist()}")
                break
        
        if not found_col:
            # Column not found - last resort: set to 0
            print("DEBUG: 'Attended % to Date.' column not found, setting attendance_pct to 0")
            attendance_normalized['attendance_pct'] = 0.0
    
    # Normalize the attendance_pct value (ensure it's 0-100 range)
    if 'attendance_pct' in attendance_normalized.columns:
        attendance_normalized['attendance_pct'] = attendance_normalized['attendance_pct'].apply(normalize_pct)
    
    # Set missed_pct to 0 since we're not using it in simplified version
    attendance_normalized['missed_pct'] = 0.0
    
    # Debug: Print sample data to verify transformations
    if 'Student Name' in attendance_normalized.columns:
        sample_cols = ['Student Name']
        if 'Attended Hours to Date' in attendance_normalized.columns:
            sample_cols.append('Attended Hours to Date')
        if 'Attended % to Date.' in attendance_normalized.columns:
            sample_cols.append('Attended % to Date.')
        print("\n=== Attendance DataFrame Sample (after normalization) ===")
        print(attendance_normalized[sample_cols].head())
        print("========================================================\n")
    
    return grades_normalized, attendance_normalized


def merge_data(grades_df: pd.DataFrame, attendance_df: pd.DataFrame) -> pd.DataFrame:
    """
    Merge grades and attendance data using INNER JOIN.
    First tries merging by Student#, then falls back to Student Name if Student# is inconsistent.
    Only includes students present in BOTH sheets.
    """
    print(f"DEBUG: merge_data called - grades_df shape: {grades_df.shape}, columns: {list(grades_df.columns)}")
    print(f"DEBUG: merge_data called - attendance_df shape: {attendance_df.shape}, columns: {list(attendance_df.columns)}")
    
    # Ensure Student Name exists in both DataFrames (required for fallback merge)
    # Check and create if missing
    if 'Student Name' not in grades_df.columns:
        print(f"ERROR: 'Student Name' not found in grades_df. Available columns: {list(grades_df.columns)}")
        if len(grades_df) > 0:
            grades_df['Student Name'] = 'Unknown'
            print("WARNING: Created placeholder 'Student Name' column in grades_df")
        else:
            grades_df['Student Name'] = pd.Series([], dtype=str)
            print("WARNING: Created empty 'Student Name' column in empty grades_df")
    
    if 'Student Name' not in attendance_df.columns:
        print(f"ERROR: 'Student Name' not found in attendance_df. Available columns: {list(attendance_df.columns)}")
        if len(attendance_df) > 0:
            attendance_df['Student Name'] = 'Unknown'
            print("WARNING: Created placeholder 'Student Name' column in attendance_df")
        else:
            attendance_df['Student Name'] = pd.Series([], dtype=str)
            print("WARNING: Created empty 'Student Name' column in empty attendance_df")
    
    # Verify columns exist now
    if 'Student Name' not in grades_df.columns:
        raise ValueError(f"'Student Name' column missing in grades DataFrame after creation attempt. Available columns: {list(grades_df.columns)}")
    
    if 'Student Name' not in attendance_df.columns:
        raise ValueError(f"'Student Name' column missing in attendance DataFrame after creation attempt. Available columns: {list(attendance_df.columns)}")
    
    # Clean Student Name for consistent matching
    try:
        # Ensure we're working with a Series, not a DataFrame
        student_name_series = safe_get_series(grades_df, 'Student Name')
        if len(student_name_series) > 0:
            grades_df['Student Name'] = student_name_series.astype(str).str.strip()
        else:
            grades_df['Student Name'] = student_name_series.astype(str)
        
        student_name_series = safe_get_series(attendance_df, 'Student Name')
        if len(student_name_series) > 0:
            attendance_df['Student Name'] = student_name_series.astype(str).str.strip()
        else:
            attendance_df['Student Name'] = student_name_series.astype(str)
        
        print("DEBUG: Successfully cleaned Student Name columns")
    except KeyError as e:
        # This is the actual error we're trying to catch
        error_msg = f"KeyError accessing 'Student Name' column: {e}"
        print(f"ERROR: {error_msg}")
        print(f"grades_df columns: {list(grades_df.columns)}")
        print(f"attendance_df columns: {list(attendance_df.columns)}")
        raise ValueError(f"{error_msg}. Grades columns: {list(grades_df.columns)}, Attendance columns: {list(attendance_df.columns)}")
    except Exception as e:
        print(f"ERROR: Failed to clean Student Name columns: {e}")
        print(f"Exception type: {type(e).__name__}")
        print(f"grades_df columns: {list(grades_df.columns)}")
        print(f"attendance_df columns: {list(attendance_df.columns)}")
        # Try fallback: use apply instead of .str
        try:
            if 'Student Name' in grades_df.columns:
                student_name_series = safe_get_series(grades_df, 'Student Name')
                if len(student_name_series) > 0:
                    grades_df['Student Name'] = student_name_series.apply(lambda x: str(x).strip() if pd.notna(x) else 'Unknown')
                else:
                    grades_df['Student Name'] = student_name_series.astype(str)
            if 'Student Name' in attendance_df.columns:
                student_name_series = safe_get_series(attendance_df, 'Student Name')
                if len(student_name_series) > 0:
                    attendance_df['Student Name'] = student_name_series.apply(lambda x: str(x).strip() if pd.notna(x) else 'Unknown')
                else:
                    attendance_df['Student Name'] = student_name_series.astype(str)
            print("SUCCESS: Used fallback method to clean Student Name columns")
        except Exception as e2:
            raise ValueError(f"Failed to process Student Name column: {e}. Fallback also failed: {e2}. Grades columns: {list(grades_df.columns)}, Attendance columns: {list(attendance_df.columns)}")
    
    # Try merging by Student# first (if both have it)
    merged = None
    merge_method = None
    
    if 'Student#' in grades_df.columns and 'Student#' in attendance_df.columns:
        # Use simple approach like user's working script
        print(f"DEBUG: Before merge - grades_df shape: {grades_df.shape}, attendance_df shape: {attendance_df.shape}")
        print(f"DEBUG: Before merge - grades_df Student# sample: {grades_df['Student#'].head(5).tolist()}")
        print(f"DEBUG: Before merge - attendance_df Student# sample: {attendance_df['Student#'].head(5).tolist()}")
        print(f"DEBUG: Before merge - grades_df Student Name sample: {grades_df['Student Name'].head(5).tolist() if 'Student Name' in grades_df.columns else 'NOT FOUND'}")
        print(f"DEBUG: Before merge - attendance_df Student Name sample: {attendance_df['Student Name'].head(5).tolist() if 'Student Name' in attendance_df.columns else 'NOT FOUND'}")
        
        # Convert Student# to string for both sheets (like user's script)
        grades_student_id = safe_get_series(grades_df, 'Student#')
        grades_df['Student#'] = grades_student_id.astype(str).str.strip()
        
        attendance_student_id = safe_get_series(attendance_df, 'Student#')
        attendance_df['Student#'] = attendance_student_id.astype(str).str.strip()
        
        # Perform INNER JOIN on Student# (like user's script)
        merged_by_id = pd.merge(
            grades_df,
            attendance_df,
            on='Student#',
            how='inner',  # INNER JOIN - only include students in both sheets
            suffixes=('_grades', '_attendance')
        )
        
        print(f"✅ Merge by Student# completed: {len(merged_by_id)} rows matched")
        print(f"DEBUG: After merge - merged columns: {list(merged_by_id.columns)}")
        print(f"DEBUG: After merge - unique Student# count: {merged_by_id['Student#'].nunique()}")
        
        # Check Student Name columns after merge
        if 'Student Name_grades' in merged_by_id.columns:
            print(f"DEBUG: Student Name_grades sample (first 5): {merged_by_id['Student Name_grades'].head(5).tolist()}")
        if 'Student Name_attendance' in merged_by_id.columns:
            print(f"DEBUG: Student Name_attendance sample (first 5): {merged_by_id['Student Name_attendance'].head(5).tolist()}")
        
        # Use the merged result
        if len(merged_by_id) > 0:
            merged = merged_by_id.copy()
            merge_method = 'Student#'
            print(f"✅ Using merge by Student# (INNER JOIN): {len(merged)} rows matched")
        else:
            print("WARNING: Student# merge resulted in 0 matches, trying Student Name merge")
    
    # Fallback to Student Name merge (or use it if Student# doesn't exist)
    if merged is None:
        # Verify Student Name exists before merging
        if 'Student Name' not in grades_df.columns:
            raise ValueError(f"Cannot merge by Student Name: column missing in grades_df. Available columns: {list(grades_df.columns)}")
        if 'Student Name' not in attendance_df.columns:
            raise ValueError(f"Cannot merge by Student Name: column missing in attendance_df. Available columns: {list(attendance_df.columns)}")
        
        print(f"Merging by Student Name: grades_df={len(grades_df)} rows, attendance_df={len(attendance_df)} rows")
        try:
            print(f"DEBUG: grades_df Student Name sample: {safe_get_series(grades_df, 'Student Name').head(5).tolist()}")
            print(f"DEBUG: attendance_df Student Name sample: {safe_get_series(attendance_df, 'Student Name').head(5).tolist()}")
        except Exception as e:
            print(f"WARNING: Could not print Student Name samples: {e}")
        
        try:
            merged = pd.merge(
                grades_df,
                attendance_df,
                on='Student Name',
                how='inner',  # Changed from 'outer' to 'inner' - only include students in both sheets
                suffixes=('_grades', '_attendance'),
                indicator=True
            )
            merge_method = 'Student Name'
            print(f"Using merge by Student Name: {len(merged)} rows")
            print(f"DEBUG: Merge by Student Name results: {merged['_merge'].value_counts().to_dict()}")
            merged = merged.drop(columns=['_merge'])
        except KeyError as e:
            error_msg = f"KeyError during Student Name merge: {e}"
            print(f"ERROR: {error_msg}")
            print(f"grades_df columns: {list(grades_df.columns)}")
            print(f"attendance_df columns: {list(attendance_df.columns)}")
            raise ValueError(f"{error_msg}. Grades columns: {list(grades_df.columns)}, Attendance columns: {list(attendance_df.columns)}")
        except Exception as e:
            error_msg = f"Error during Student Name merge: {e}"
            print(f"ERROR: {error_msg}")
            print(f"Exception type: {type(e).__name__}")
            print(f"grades_df columns: {list(grades_df.columns)}")
            print(f"attendance_df columns: {list(attendance_df.columns)}")
            raise ValueError(f"{error_msg}. Grades columns: {list(grades_df.columns)}, Attendance columns: {list(attendance_df.columns)}")
    
    # Ensure Student# column exists (create from either side if needed)
    if 'Student#' not in merged.columns:
        if 'Student#_grades' in merged.columns:
            merged['Student#'] = merged['Student#_grades'].fillna(merged.get('Student#_attendance', pd.Series([None] * len(merged))))
            merged = merged.drop(columns=['Student#_grades', 'Student#_attendance'], errors='ignore')
        elif 'Student#_attendance' in merged.columns:
            merged['Student#'] = merged['Student#_attendance']
            merged = merged.drop(columns=['Student#_attendance'], errors='ignore')
        else:
            # Generate sequential IDs if neither exists
            merged['Student#'] = range(1, len(merged) + 1)
    
    # Verify merged DataFrame has Student Name
    if 'Student Name' not in merged.columns and 'Student Name_grades' not in merged.columns and 'Student Name_attendance' not in merged.columns:
        print(f"WARNING: 'Student Name' not found in merged DataFrame. Available columns: {list(merged.columns)}")
        merged['Student Name'] = 'Unknown'
    
    try:
        student_name_col = 'Student Name' if 'Student Name' in merged.columns else ('Student Name_grades' if 'Student Name_grades' in merged.columns else 'Student Name_attendance')
        print(f"After merge ({merge_method}): {len(merged)} rows, {safe_get_series(merged, student_name_col).nunique()} unique students")
        print(f"DEBUG: merged columns: {list(merged.columns)}")
        print(f"DEBUG: merged Student Name sample: {safe_get_series(merged, student_name_col).head(10).tolist()}")
    except Exception as e:
        print(f"WARNING: Could not print merge statistics: {e}")
        print(f"After merge ({merge_method}): {len(merged)} rows")
        print(f"DEBUG: merged columns: {list(merged.columns)}")
    
    # Handle Student Name - prefer grades, fallback to attendance
    # CRITICAL: Both sheets have Student Name, so after merge we get Student Name_grades and Student Name_attendance
    # We need to consolidate them properly, prioritizing grades but ensuring we don't lose any actual names
    print(f"\n=== DEBUG: Consolidating Student Name ===")
    print(f"Columns before consolidation: {[col for col in merged.columns if 'Name' in col or 'Student' in col]}")
    
    if 'Student Name_grades' in merged.columns and 'Student Name_attendance' in merged.columns:
        print(f"Both Student Name columns exist - consolidating...")
        print(f"Student Name_grades sample: {merged['Student Name_grades'].head(5).tolist()}")
        print(f"Student Name_attendance sample: {merged['Student Name_attendance'].head(5).tolist()}")
        
        # Both exist - prefer grades, but use attendance if grades is empty/null/NaN
        # Convert to string first to handle any type issues
        merged['Student Name_grades'] = merged['Student Name_grades'].astype(str).str.strip()
        merged['Student Name_attendance'] = merged['Student Name_attendance'].astype(str).str.strip()
        
        # Replace 'nan', 'None', empty strings with actual NaN for fillna to work
        merged['Student Name_grades'] = merged['Student Name_grades'].replace(['nan', 'None', ''], pd.NA)
        merged['Student Name_attendance'] = merged['Student Name_attendance'].replace(['nan', 'None', ''], pd.NA)
        
        # Consolidate: prefer grades, fallback to attendance
        merged['Student Name'] = merged['Student Name_grades'].fillna(merged['Student Name_attendance'])
        
        # If still NaN, use the other column
        mask = merged['Student Name'].isna()
        merged.loc[mask, 'Student Name'] = merged.loc[mask, 'Student Name_attendance']
        
        # Convert back to string and clean up
        merged['Student Name'] = merged['Student Name'].astype(str).str.strip()
        
        # Only set to 'Unknown' if BOTH columns were empty
        both_empty = (merged['Student Name_grades'].isna() | (merged['Student Name_grades'] == '')) & \
                     (merged['Student Name_attendance'].isna() | (merged['Student Name_attendance'] == ''))
        merged.loc[both_empty, 'Student Name'] = 'Unknown'
        
        merged = merged.drop(columns=['Student Name_grades', 'Student Name_attendance'])
        print(f"✅ Consolidated Student Name - sample: {merged['Student Name'].head(5).tolist()}")
        print(f"   Valid names: {(merged['Student Name'] != 'Unknown').sum()}, Unknown: {(merged['Student Name'] == 'Unknown').sum()}")
        
    elif 'Student Name_grades' in merged.columns:
        print(f"Only Student Name_grades exists - using it...")
        merged['Student Name'] = merged['Student Name_grades'].astype(str).str.strip()
        merged['Student Name'] = merged['Student Name'].replace(['nan', 'None', ''], 'Unknown')
        merged = merged.drop(columns=['Student Name_grades'])
        print(f"✅ Using Student Name from grades - sample: {merged['Student Name'].head(5).tolist()}")
        
    elif 'Student Name_attendance' in merged.columns:
        print(f"Only Student Name_attendance exists - using it...")
        merged['Student Name'] = merged['Student Name_attendance'].astype(str).str.strip()
        merged['Student Name'] = merged['Student Name'].replace(['nan', 'None', ''], 'Unknown')
        merged = merged.drop(columns=['Student Name_attendance'])
        print(f"✅ Using Student Name from attendance - sample: {merged['Student Name'].head(5).tolist()}")
        
    elif 'Student Name' in merged.columns:
        # Already consolidated, but clean it up
        print(f"Student Name already exists - cleaning it up...")
        merged['Student Name'] = merged['Student Name'].astype(str).str.strip()
        merged['Student Name'] = merged['Student Name'].replace(['nan', 'None', ''], 'Unknown')
        print(f"✅ Student Name already exists - sample: {merged['Student Name'].head(5).tolist()}")
        
    else:
        # Last resort: create placeholder, but log warning
        print(f"ERROR: No Student Name column found after merge! Creating placeholder.")
        print(f"Available columns: {list(merged.columns)}")
        merged['Student Name'] = 'Unknown'
    
    print(f"=== END DEBUG: Student Name consolidation ===\n")
    
    # Handle Program Name - prefer grades, fallback to attendance
    if 'Program Name' in merged.columns:
        pass  # Already from grades
    elif 'Program Name_grades' in merged.columns:
        merged['Program Name'] = merged['Program Name_grades']
        merged = merged.drop(columns=['Program Name_grades'])
    elif 'Program Name_attendance' in merged.columns:
        merged['Program Name'] = merged['Program Name_attendance']
        merged = merged.drop(columns=['Program Name_attendance'])
    else:
        merged['Program Name'] = 'Unknown'
    
    # Preserve Campus Login URL from attendance sheet (if it exists)
    if 'Campus Login URL' not in merged.columns:
        # Check if it exists with suffix
        if 'Campus Login URL_attendance' in merged.columns:
            merged['Campus Login URL'] = merged['Campus Login URL_attendance']
            merged = merged.drop(columns=['Campus Login URL_attendance'])
        else:
            merged['Campus Login URL'] = None
    
    # Track which data came from which sheet BEFORE merging columns
    # This helps us determine if data is actually missing vs. just 0.0
    has_grade_from_grades = 'grade_pct_grades' in merged.columns
    has_grade_from_attendance = 'grade_pct_attendance' in merged.columns
    has_attendance_from_grades = 'attendance_pct_grades' in merged.columns
    has_attendance_from_attendance = 'attendance_pct_attendance' in merged.columns
    
    # Preserve attendance_pct from attendance sheet
    # Track if data actually exists (not just if it's 0.0)
    if 'attendance_pct' not in merged.columns:
        if 'attendance_pct_attendance' in merged.columns:
            # Check if data exists BEFORE filling NaN
            merged['_has_attendance'] = merged['attendance_pct_attendance'].notna()
            merged['attendance_pct'] = merged['attendance_pct_attendance'].fillna(0.0)
            merged = merged.drop(columns=['attendance_pct_attendance'])
        elif 'attendance_pct_grades' in merged.columns:
            merged['_has_attendance'] = merged['attendance_pct_grades'].notna()
            merged['attendance_pct'] = merged['attendance_pct_grades'].fillna(0.0)
            merged = merged.drop(columns=['attendance_pct_grades'])
        else:
            # Try to find any attendance percentage column
            found_col = None
            for col in merged.columns:
                if 'attended' in str(col).lower() and ('%' in str(col) or 'pct' in str(col).lower()):
                    found_col = col
                    merged['_has_attendance'] = merged[col].notna()
                    merged['attendance_pct'] = merged[col].fillna(0.0)
                    break
            if not found_col:
                merged['attendance_pct'] = 0.0
                merged['_has_attendance'] = False
    else:
        # If attendance_pct already exists, check if it came from either sheet
        merged['_has_attendance'] = (
            (merged.get('attendance_pct_attendance', pd.Series([False] * len(merged))).notna()) |
            (merged.get('attendance_pct_grades', pd.Series([False] * len(merged))).notna()) |
            merged['attendance_pct'].notna()
        )
        merged['attendance_pct'] = merged['attendance_pct'].fillna(0.0)
    
    # Preserve grade_pct from grades sheet
    if 'grade_pct' not in merged.columns:
        if 'grade_pct_grades' in merged.columns:
            merged['_has_grade'] = merged['grade_pct_grades'].notna()
            merged['grade_pct'] = merged['grade_pct_grades'].fillna(0.0)
            merged = merged.drop(columns=['grade_pct_grades'])
        elif 'grade_pct_attendance' in merged.columns:
            merged['_has_grade'] = merged['grade_pct_attendance'].notna()
            merged['grade_pct'] = merged['grade_pct_attendance'].fillna(0.0)
            merged = merged.drop(columns=['grade_pct_attendance'])
        else:
            merged['grade_pct'] = 0.0
            merged['_has_grade'] = False
    else:
        # If grade_pct already exists, check if it came from either sheet
        merged['_has_grade'] = (
            (merged.get('grade_pct_grades', pd.Series([False] * len(merged))).notna()) |
            (merged.get('grade_pct_attendance', pd.Series([False] * len(merged))).notna()) |
            merged['grade_pct'].notna()
        )
        merged['grade_pct'] = merged['grade_pct'].fillna(0.0)
    
    # Determine data status - label missing data based on whether data actually exists
    # Use the _has_grade and _has_attendance flags to determine if data is missing
    grade_missing = ~merged['_has_grade']
    attendance_missing = ~merged['_has_attendance']
    
    # Create data_status column
    merged['data_status'] = np.where(
        grade_missing & attendance_missing,
        'Missing Both',
        np.where(
            grade_missing,
            'Missing Grade',
            np.where(
                attendance_missing,
                'Missing Attendance',
                'Complete'
            )
        )
    )
    
    print(f"Data status counts: {merged['data_status'].value_counts().to_dict()}")
    
    # Remove helper columns
    if '_has_grade' in merged.columns:
        merged = merged.drop(columns=['_has_grade'])
    if '_has_attendance' in merged.columns:
        merged = merged.drop(columns=['_has_attendance'])
    
    # Remove duplicates by Student# (keep first occurrence)
    # This handles cases where a student appears multiple times
    merged = merged.drop_duplicates(subset=['Student#'], keep='first')
    
    print(f"After deduplication: {len(merged)} rows")
    print(f"DEBUG: Final merged shape: {merged.shape}")
    print(f"DEBUG: Final merged Student# count: {merged['Student#'].nunique()}")
    
    return merged

